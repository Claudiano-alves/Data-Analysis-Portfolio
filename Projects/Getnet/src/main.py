import pandas as pd
from openpyxl import load_workbook
from pipeline_funil import executar_pipeline_funil

def atualizar_arquivo_funil(caminho_destino=None):
    """
    Executa o pipeline do funil e atualiza o arquivo Excel com novos dados.
    
    Args:
        caminho_destino (str, optional): Caminho do arquivo Excel de destino.
            Se None, usa o caminho padrão.
    
    Returns:
        dict: Dicionário com informações sobre a execução:
            - 'sucesso': bool indicando se a operação foi bem-sucedida
            - 'mensagem': str com mensagem descritiva
            - 'registros_novos': int com quantidade de registros adicionados
            - 'ultima_data': str com a última data processada
    """
    
    # Define o caminho padrão se não for fornecido
    if caminho_destino is None:
        caminho_destino = r"\\trc-dc-ad\Planejamento\MIS\CARTEIRAS\GetNet\df_csvBI_padronizado.xlsx"
    
    resultado = {
        'sucesso': False,
        'mensagem': '',
        'registros_novos': 0,
        'ultima_data': None
    }
    
    try:
        # 1. Executa o pipeline do funil
        print("🔄 Executando pipeline do funil...")
        df_funil_final = executar_pipeline_funil()
        
        # Renomeia 'DATA' para 'data' no df_funil_final para padronizar
        if 'DATA' in df_funil_final.columns:
            df_funil_final.rename(columns={'DATA': 'data'}, inplace=True)
        
        df_funil_final['data'] = pd.to_datetime(df_funil_final['data'], errors='coerce')
        
        try:
            # 2. Lê o arquivo existente da aba 'df_csvBI_padronizado'
            print("📂 Lendo arquivo existente...")
            df_existente = pd.read_excel(caminho_destino, sheet_name='df_csvBI_padronizado')
            
            # 3. Converte a coluna 'data' para datetime
            df_existente['data'] = pd.to_datetime(df_existente['data'], errors='coerce')
            
            # 4. Obtém a última data do arquivo existente
            ultima_data_existente = df_existente['data'].max()
            
            print(f"📅 Última data no arquivo: {ultima_data_existente.strftime('%Y-%m-%d')}")
            resultado['ultima_data'] = ultima_data_existente.strftime('%Y-%m-%d')
            
            # 5. Filtra apenas dados NOVOS (posteriores à última data existente)
            df_novos = df_funil_final[df_funil_final['data'] > ultima_data_existente]
            
            if len(df_novos) == 0:
                resultado['sucesso'] = True
                resultado['mensagem'] = "Nenhum dado novo para adicionar. Arquivo não foi atualizado."
                resultado['registros_novos'] = 0
                print("⚠️ " + resultado['mensagem'])
                return resultado
            
            print(f"✅ {len(df_novos)} novos registros serão adicionados")
            
            # 6. Carrega o workbook existente SEM destruir formatação
            wb = load_workbook(caminho_destino)
            ws = wb['df_csvBI_padronizado']
            
            # 7. Encontra a próxima linha vazia (após a última linha com dados)
            proxima_linha = ws.max_row + 1
            linha_inicial = proxima_linha
            
            # 8. Adiciona apenas os DADOS NOVOS (sem cabeçalho)
            for _, row in df_novos.iterrows():
                for col_idx, valor in enumerate(row, start=1):
                    ws.cell(row=proxima_linha, column=col_idx, value=valor)
                proxima_linha += 1
            
            # 9. Salva o arquivo preservando TUDO (formatação, abas, tabelas dinâmicas)
            wb.save(caminho_destino)
            wb.close()
            
            resultado['sucesso'] = True
            resultado['mensagem'] = f"Arquivo atualizado com sucesso! {len(df_novos)} registros adicionados."
            resultado['registros_novos'] = len(df_novos)
            
            print(f"💾 {resultado['mensagem']}")
            print(f"📊 Registros adicionados a partir da linha {linha_inicial}")
            
        except FileNotFoundError:
            # Arquivo não existe, cria um novo
            print("⚠️ Arquivo não encontrado. Criando novo arquivo com os dados gerados.")
            with pd.ExcelWriter(caminho_destino, engine='openpyxl') as writer:
                df_funil_final.to_excel(writer, sheet_name='df_csvBI_padronizado', index=False)
            
            resultado['sucesso'] = True
            resultado['mensagem'] = "Arquivo criado com sucesso!"
            resultado['registros_novos'] = len(df_funil_final)
            resultado['ultima_data'] = df_funil_final['data'].max().strftime('%Y-%m-%d')
            
            print(f"💾 {resultado['mensagem']}")
        
        return resultado
        
    except Exception as e:
        resultado['sucesso'] = False
        resultado['mensagem'] = f"Erro ao processar arquivo: {str(e)}"
        print(f"❌ {resultado['mensagem']}")
        raise

# Exemplo de uso:
if __name__ == "__main__":
    resultado = atualizar_arquivo_funil()
    print("\n📋 Resultado da execução:")
    print(f"   Sucesso: {resultado['sucesso']}")
    print(f"   Mensagem: {resultado['mensagem']}")
    print(f"   Novos registros: {resultado['registros_novos']}")
    print(f"   Última data: {resultado['ultima_data']}")