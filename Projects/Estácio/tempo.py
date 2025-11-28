from datetime import datetime, timedelta
import pandas as pd
from db_connection import get_connection
from querie import get_query_tempos
from openpyxl import load_workbook

base_path = r"\\trc-dc-ad\OperacionalEstacio\PA provida\PA Provida - MODELO.xlsx"

def obter_range_datas_xlsx(base_path: str):
    """
    Lê arquivo XLSX e retorna um range de datas:
    - data_inicio: dia seguinte à última data encontrada na coluna DATA_LOGIN
    - data_fim: dia anterior ao dia atual (hoje - 1)
    
    Retorna: (data_inicio, data_fim) no formato 'YYYY-MM-DD'
    """
    # Data fim: ontem (hoje - 1)
    hoje = datetime.today()
    data_fim = (hoje - timedelta(days=1)).strftime("%Y-%m-%d")
    
    ultima_data = None
    
    try:
        # Abre o arquivo XLSX
        wb = load_workbook(base_path, data_only=True)
        
        # Tenta acessar a aba 'Analítico'
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() == 'analítico' or name.lower() == 'analitico':
                sheet_name = name
                break
        
        if sheet_name is None:
            print("Aba 'Analítico' não encontrada")
            mes_atual = hoje.strftime("%Y%m")
            data_inicio = datetime.strptime(mes_atual + "01", "%Y%m%d").strftime("%Y-%m-%d")
            return data_inicio, data_fim
        
        ws = wb[sheet_name]
        
        # Procura a coluna DATA_LOGIN no cabeçalho
        indice_coluna_data = None
        for idx, cell in enumerate(ws[1], start=1):
            if cell.value and isinstance(cell.value, str):
                if cell.value.strip().upper() == 'DATA_LOGIN':
                    indice_coluna_data = idx
                    break
        
        if indice_coluna_data is None:
            print("Coluna 'DATA_LOGIN' não encontrada no cabeçalho")
            mes_atual = hoje.strftime("%Y%m")
            data_inicio = datetime.strptime(mes_atual + "01", "%Y%m%d").strftime("%Y-%m-%d")
            return data_inicio, data_fim
        
        # Lê apenas a coluna DATA_LOGIN (pulando o cabeçalho)
        for row in ws.iter_rows(min_row=2, min_col=indice_coluna_data, max_col=indice_coluna_data):
            cell = row[0]
            
            if cell.value is not None:
                try:
                    # Se já for datetime
                    if isinstance(cell.value, datetime):
                        data = cell.value
                    # Se for string
                    elif isinstance(cell.value, str):
                        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y%m%d", "%d-%m-%Y"]:
                            try:
                                data = datetime.strptime(cell.value, fmt)
                                break
                            except ValueError:
                                continue
                        else:
                            continue
                    else:
                        continue
                    
                    # Atualiza última_data se for mais recente
                    if ultima_data is None or data > ultima_data:
                        ultima_data = data
                except (ValueError, TypeError):
                    continue
        
        wb.close()
        
    except Exception as e:
        print(f"Erro ao processar arquivo: {e}")
    
    # Define data de início
    if ultima_data is None:
        mes_atual = hoje.strftime("%Y%m")
        data_inicio = datetime.strptime(mes_atual + "01", "%Y%m%d").strftime("%Y-%m-%d")
    else:
        data_inicio = (ultima_data + timedelta(days=1)).strftime("%Y-%m-%d")
        print(f"Última data encontrada no arquivo: {ultima_data.strftime('%d/%m/%Y')}")
    
    return data_inicio, data_fim

def gerar_dados_analitico():
    """
    Gera dados diários e adiciona na aba 'Analítico' do arquivo XLSX.
    Preserva toda a formatação e apenas adiciona novas linhas ao final.
    """
    # Obtém o range de datas
    data_inicio, data_fim = obter_range_datas_xlsx(base_path)
    
    # Converter datas para objetos datetime
    dt_inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
    dt_fim = datetime.strptime(data_fim, "%Y-%m-%d")
    
    # Validação: se data_inicio > data_fim, não há nada a fazer
    if dt_inicio > dt_fim:
        print(f"Nenhum dado a gerar. Dados já estão atualizados até: {(dt_inicio - timedelta(days=1)).strftime('%d/%m/%Y')}")
        return
    
    print(f"Gerando dados de {dt_inicio.strftime('%d/%m/%Y')} até {dt_fim.strftime('%d/%m/%Y')}")
    
    # Conexão com o banco
    conn_bd2 = get_connection("TRC-DC-BD2", "PLANEJAMENTO")
    
    # Consulta SQL para todo o período
    query = get_query_tempos(data_inicio, data_fim)
    df_novos_dados = pd.read_sql(query, conn_bd2)
    conn_bd2.close()
    
    if df_novos_dados.empty:
        print("Nenhum dado retornado da consulta.")
        return
    
    print(f"Total de {len(df_novos_dados)} registros obtidos do banco de dados.")
    
    try:
        # Carrega o workbook preservando formatação
        wb = load_workbook(base_path)
        
        # Encontra a aba Analítico
        sheet_name = None
        for name in wb.sheetnames:
            if name.lower() == 'analítico' or name.lower() == 'analitico':
                sheet_name = name
                break
        
        if sheet_name is None:
            print("Erro: Aba 'Analítico' não encontrada")
            return
        
        ws = wb[sheet_name]
        
        # Encontra a última linha com dados
        ultima_linha = ws.max_row
        
        print(f"Adicionando dados a partir da linha {ultima_linha + 1}...")
        
        # Adiciona os novos dados linha por linha
        for idx, row in df_novos_dados.iterrows():
            linha_destino = ultima_linha + idx + 1
            for col_idx, valor in enumerate(row, start=1):
                ws.cell(row=linha_destino, column=col_idx, value=valor)
        
        # Salva o arquivo preservando formatação
        wb.save(base_path)
        wb.close()
        
        print(f"✓ {len(df_novos_dados)} registros adicionados com sucesso!")
        print(f"✓ Arquivo atualizado: {base_path}")
        
    except Exception as e:
        print(f"Erro ao processar arquivo: {e}")
        import traceback
        traceback.print_exc()


# Executar a função
# Por padrão, gera até ontem
gerar_dados_analitico()

# Ou você pode especificar uma data_fim customizada:
# gerar_arquivos_tempos(data_fim="2025-11-20")
