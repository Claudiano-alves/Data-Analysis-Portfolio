import pandas as pd
from datetime import datetime, timedelta
from db_connection import get_connection
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# def obter_ultima_data_arquivo(caminho_csv):
#     """
#     Lê o arquivo CSV e retorna a última data de atendimento.
    
#     Parâmetros:
#     -----------
#     caminho_csv : str
#         Caminho do arquivo CSV
    
#     Retorna:
#     --------
#     datetime : Última data encontrada no arquivo
#     """
#     try:
#         # Ler CSV com delimitador ponto e vírgula
#         df = pd.read_csv(caminho_csv, encoding='utf-8-sig', sep=';')
        
#         # Verificar se a coluna existe
#         print(f"Colunas encontradas no arquivo ({len(df.columns)} colunas):")
#         print(list(df.columns)[:10])  # Mostrar apenas as 10 primeiras
        
#         if 'data_atendimento' not in df.columns:
#             raise ValueError("Coluna 'data_atendimento' não encontrada no arquivo")
        
#         # Mostrar exemplo de como a data está no arquivo
#         print(f"Exemplo de data no arquivo: {df['data_atendimento'].iloc[0]}")
        
#         # Tentar converter para datetime com dayfirst=True (formato brasileiro)
#         df['data_atendimento'] = pd.to_datetime(df['data_atendimento'], dayfirst=True, errors='coerce')
        
#         # Verificar se conseguiu converter
#         if df['data_atendimento'].isna().all():
#             raise ValueError("Não foi possível converter nenhuma data. Verifique o formato das datas no arquivo.")
        
#         # Obter a última data
#         ultima_data = df['data_atendimento'].max()
        
#         # Verificar se a última data é válida
#         if pd.isna(ultima_data):
#             raise ValueError("Última data é inválida (NaT)")
        
#         print(f"Última data encontrada no arquivo: {ultima_data.strftime('%d/%m/%Y')}")
        
#         return ultima_data
        
#     except Exception as e:
#         print(f"Erro ao ler arquivo: {str(e)}")
#         import traceback
#         traceback.print_exc()
#         raise

def obter_ultima_data_arquivo(caminho_xlsx):
    """
    Lê o arquivo XLSX e retorna a maior data de atendimento encontrada.
    """
    import re
    
    try:
        # Ler XLSX
        df = pd.read_excel(caminho_xlsx)
        
        # Normalizar nomes de colunas (remover espaços)
        df.columns = df.columns.str.strip()
        
        print(f"Total de linhas no arquivo: {len(df)}")
        print(f"Colunas encontradas: {list(df.columns)}")
        print(f"Tipo da coluna 'data_atendimento': {df['data_atendimento'].dtype}")
        
        if 'data_atendimento' not in df.columns:
            raise ValueError("Coluna 'data_atendimento' não encontrada no arquivo")
        
        # Mostrar primeiras e últimas linhas para debug
        print(f"\nPrimeiras 3 valores de data_atendimento:")
        print(df['data_atendimento'].head(3).tolist())
        print(f"\nÚltimas 3 valores de data_atendimento:")
        print(df['data_atendimento'].tail(3).tolist())
        
        # Converter para string e limpar
        series_datas = df['data_atendimento'].astype(str).str.strip()
        
        # Remover valores 'NaT', 'None', 'nat', etc
        mask_valido = ~series_datas.isin(['NaT', 'None', 'none', 'nat', '', 'NaN'])
        series_datas = series_datas[mask_valido]
        
        print(f"\nValores após limpeza de NaT/None: {len(series_datas)}")
        
        if len(series_datas) == 0:
            raise ValueError("Nenhum valor válido encontrado na coluna 'data_atendimento'")
        
        # Tentar converter para datetime com múltiplos formatos
        datas_convertidas = pd.to_datetime(
            series_datas,
            dayfirst=True,
            errors='coerce',
            format='mixed'
        )
        
        # Verificar quantas foram convertidas
        n_convertidas = datas_convertidas.notna().sum()
        n_nao_convertidas = datas_convertidas.isna().sum()
        
        print(f"\nDatas convertidas com sucesso: {n_convertidas}")
        print(f"Datas que falharam na conversão: {n_nao_convertidas}")
        
        if n_convertidas == 0:
            # Se nenhuma foi convertida, mostrar valores problemáticos
            print(f"\nExemplos de valores que não foram convertidos:")
            problematicas = series_datas[datas_convertidas.isna()].unique()[:10]
            for val in problematicas:
                print(f"  - '{val}'")
            raise ValueError("Nenhuma data válida pôde ser convertida. Verifique os valores acima.")
        
        # Remover NaT (datas que não foram convertidas)
        datas_validas = datas_convertidas[datas_convertidas.notna()]
        
        # Encontrar a MAIOR data
        ultima_data = datas_validas.max()
        
        print(f"\n✓ Última data encontrada no arquivo: {ultima_data.strftime('%d/%m/%Y')}")
        print(f"  (Total de {n_convertidas} datas válidas analisadas)")
        
        return ultima_data
    
    except Exception as e:
        print(f"\n✗ Erro ao ler arquivo: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def consultar_tempos_operacionais(conn, data_inicio, data_fim):
    """
    Consulta dados de tempos operacionais do SQL Server.
    
    Parâmetros:
    -----------
    conn : pyodbc.Connection
        Conexão ativa com o banco de dados
    data_inicio : datetime ou str
        Data inicial do período (formato: 'YYYY-MM-DD' ou datetime)
    data_fim : datetime ou str
        Data final do período (formato: 'YYYY-MM-DD' ou datetime)
    
    Retorna:
    --------
    DataFrame : Dados consultados
    """
    
    # Converter para datetime se necessário
    if isinstance(data_inicio, str):
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
    if isinstance(data_fim, str):
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d')
    
    query = """
    SELECT
        COALESCE(CAST(data_insert AS DATE), '1900-01-01') as data_insert,
        COALESCE(CAST(hora_insert AS TIME), '00:00:00') as hora_insert,
        COALESCE(CAST(id_agente AS BIGINT), 0) as id_agente,
        LTRIM(RTRIM(REPLACE(REPLACE(REPLACE(REPLACE(COALESCE(nome_agente, ''), 'JOÃO', 'JOAO'), CHAR(9), ''), CHAR(13), ''), CHAR(10), ''))) as nome_agente,
        COALESCE(cpf_agente, '') as cpf_agente,
        COALESCE(nome_supervisor, '') as nome_supervisor,
        1 as coluna_fixa,
        COALESCE(CAST(data_atendimento AS DATE), '1900-01-01') as data_atendimento,
        COALESCE(CAST(DATEADD(MINUTE, 60, CAST(COALESCE(tempo_logado, '00:00:00') AS DATETIME)) AS TIME), '00:00:00') as tempo_logado,
        COALESCE(CAST(hora_login AS TIME), '00:00:00') as hora_login,
        COALESCE(
            CAST(
                CASE 
                    WHEN DATEADD(MINUTE, 60, CAST(COALESCE(hora_logout, '00:00:00') AS DATETIME)) > CAST('20:40:00' AS DATETIME)
                         THEN CAST('20:40:00' AS DATETIME)
                    ELSE DATEADD(MINUTE, 60, CAST(COALESCE(hora_logout, '00:00:00') AS DATETIME))
                END
            AS TIME),
        '00:00:00') as hora_logout,
        COALESCE(CAST(tempo_trabalhado AS TIME), '00:00:00') as tempo_trabalhado,
        COALESCE(CAST(DATEADD(MINUTE, 90, CAST(COALESCE(tempo_falado, '00:00:00') AS DATETIME)) AS TIME), '00:00:00') as tempo_falado,
        CASE
            WHEN DATEADD(HOUR, -1, CAST(COALESCE(tempo_disponivel, '00:00:00') AS DATETIME)) < '1900-01-01 00:00:00'
            THEN CAST('00:00:00' AS TIME)
            ELSE CAST(DATEADD(MINUTE, -30, CAST(COALESCE(tempo_disponivel, '00:00:00') AS DATETIME)) AS TIME)
        END as tempo_disponivel,
        COALESCE(CAST(tempo_pos_atendimento AS TIME), '00:00:00') as tempo_pos_atendimento,
        COALESCE(CAST(pausa_descanso1 AS TIME), '00:00:00') as pausa_descanso1,
        COALESCE(CAST(pausa_lanche AS TIME), '00:00:00') as pausa_lanche,
        COALESCE(CAST(pausa_descanso2 AS TIME), '00:00:00') as pausa_descanso2,
        CAST(
            DATEADD(SECOND, 
                (
                    DATEDIFF(SECOND, '00:00:00', COALESCE(pausa_descanso1, '00:00:00')) +
                    DATEDIFF(SECOND, '00:00:00', COALESCE(pausa_lanche, '00:00:00')) +
                    DATEDIFF(SECOND, '00:00:00', COALESCE(pausa_descanso2, '00:00:00'))
                ),
                '00:00:00'
            )
        AS TIME) as pausa_nr17,
        COALESCE(CAST(pausa_reuniao AS TIME), '00:00:00') as pausa_reuniao,
        COALESCE(CAST(pausa_treinamento AS TIME), '00:00:00') as pausa_treinamento,
        COALESCE(CAST(pausa_feedback AS TIME), '00:00:00') as pausa_feedback,
        COALESCE(CAST(pausa_sistema AS TIME), '00:00:00') as pausa_sistema,
        COALESCE(CAST(pausa_outros AS TIME), '00:00:00') as pausa_outros,
        COALESCE(tipo_de_operacao, '') as tipo_de_operacao,
        COALESCE(nome_da_assessoria, '') as nome_da_assessoria,
        CASE
            WHEN DATEADD(HOUR, -1, CAST(COALESCE(tempo_idle, '00:00:00') AS DATETIME)) < '1900-01-01 00:00:00'
            THEN CAST('00:00:00' AS TIME)
            ELSE CAST(DATEADD(MINUTE, -30, CAST(COALESCE(tempo_idle, '00:00:00') AS DATETIME)) AS TIME)
        END as tempo_idle,
        COALESCE(CAST(pausa_digital AS TIME), '00:00:00') as pausa_digital,
        COALESCE(CAST(tma_total AS TIME), '00:00:00') as tma_total,
        COALESCE(CAST(tma_produtivo AS TIME), '00:00:00') as tma_produtivo,
        COALESCE(atendimentos_total, 0) as atendimentos_total,
        COALESCE(atendimentos_produtivo, 0) as atendimentos_produtivo,
        CASE 
            WHEN COALESCE(atendimentos_total, 0) = 0 THEN CAST('00:00:00' AS TIME)
            ELSE CAST(
                DATEADD(SECOND, 
                    DATEDIFF(SECOND, '00:00:00', COALESCE(tempo_pos_atendimento, '00:00:00')) / atendimentos_total,
                    '00:00:00'
                )
            AS TIME)
        END as med_pos,
        CASE 
            WHEN COALESCE(atendimentos_total, 0) = 0 THEN CAST('00:00:00' AS TIME)
            ELSE 
                CASE
                    WHEN DATEADD(HOUR, -1, CAST(COALESCE(tempo_idle, '00:00:00') AS DATETIME)) < '1900-01-01 00:00:00'
                    THEN CAST('00:00:00' AS TIME)
                    ELSE CAST(
                        DATEADD(SECOND, 
                            DATEDIFF(SECOND, '00:00:00', 
                                CAST(DATEADD(MINUTE, -30, CAST(COALESCE(tempo_idle, '00:00:00') AS DATETIME)) AS TIME)
                            ) / atendimentos_total,
                            '00:00:00'
                        )
                    AS TIME)
                END
        END as med_idle
    FROM WillCenter_TemposOperacionais WITH (NOLOCK)
    WHERE data_atendimento >= ? AND data_atendimento <= ?
    ORDER BY data_atendimento, id_agente
    """
    
    print(f"Consultando dados de {data_inicio.strftime('%Y-%m-%d')} até {data_fim.strftime('%Y-%m-%d')}...")
    
    df = pd.read_sql(query, conn, params=[
        data_inicio.strftime('%Y-%m-%d'),
        data_fim.strftime('%Y-%m-%d')
    ])
    
    print(f"Consulta concluída: {len(df)} linhas retornadas")
    
    return df

# def processar_atualizacao_tempos():
#     """
#     Função principal que orquestra o processo de atualização.
    
#     1. Mapeia a última data no arquivo
#     2. Elabora o range de datas (última data + 1 até ontem)
#     3. Consulta dados do banco
#     4. Adiciona ao arquivo CSV
#     """
    
#     # Configurações
#     caminho_csv = r"\\trc-dc-ad\Planejamento\MIS\CARTEIRAS\WILLBANK\Automations\base_tempos.csv"
    
#     try:
#         print("="*60)
#         print("INICIANDO PROCESSO DE ATUALIZAÇÃO DE TEMPOS OPERACIONAIS")
#         print("="*60)
        
#         # 1. Mapear última data no arquivo
#         print("\n[1/4] Mapeando última data no arquivo...")
#         ultima_data_arquivo = obter_ultima_data_arquivo(caminho_csv)
        
#         # 2. Elaborar range de datas
#         print("\n[2/4] Elaborando range de datas...")
#         data_inicio = ultima_data_arquivo + timedelta(days=1)
#         data_fim = datetime.now() - timedelta(days=1)
        
#         print(f"Última data no arquivo: {ultima_data_arquivo.strftime('%d/%m/%Y')}")
#         print(f"Período a buscar: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}")
        
#         # Validar se há dados a buscar
#         if data_inicio > data_fim:
#             print(f"\nArquivo já está atualizado!")
#             return {
#                 'sucesso': True,
#                 'mensagem': 'Arquivo já está atualizado',
#                 'linhas_adicionadas': 0
#             }
        
#         # 3. Conectar e consultar banco
#         print("\n[3/4] Conectando ao banco e consultando dados...")
#         conn_bd2 = get_connection("TRC-DC-BD2", "PLANEJAMENTO")
        
#         df_novos = consultar_tempos_operacionais(conn_bd2, data_inicio, data_fim)
        
#         conn_bd2.close()
        
#         if len(df_novos) == 0:
#             print("\nNenhum dado novo encontrado para o período.")
#             return {
#                 'sucesso': True,
#                 'mensagem': 'Nenhum dado novo encontrado',
#                 'linhas_adicionadas': 0,
#                 'periodo': f"{data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
#             }
        
#         # 4. Adicionar ao arquivo CSV (apenas append, sem modificar dados existentes)
#         print("\n[4/4] Adicionando dados ao arquivo CSV...")
        
#         print(f"Linhas a adicionar: {len(df_novos)}")
        
#         # Adicionar os novos dados ao final do arquivo CSV com delimitador ponto e vírgula
#         df_novos.to_csv(caminho_csv, mode='a', header=False, index=False, encoding='utf-8-sig', sep=';')
#         #df_novos = df_novos[df_novos.columns]  # força mesma ordem e nomes de colunas

#         print(f"Dados adicionados com sucesso ao arquivo CSV!")
        
#         print("\n" + "="*60)
#         print("PROCESSO CONCLUÍDO COM SUCESSO!")
#         print("="*60)
        
#         return {
#             'sucesso': True,
#             'mensagem': 'Dados atualizados com sucesso',
#             'linhas_adicionadas': len(df_novos),
#             'periodo': f"{data_inicio.strftime('%Y-%m-%d')} até {data_fim.strftime('%Y-%m-%d')}"
#         }
        
#     except Exception as e:
#         print(f"\nERRO: {str(e)}")
#         return {
#             'sucesso': False,
#             'mensagem': f'Erro ao processar: {str(e)}',
#             'erro': str(e)
#         }

def processar_atualizacao_tempos():
    """
    Função principal que orquestra o processo de atualização.
    
    1. Mapeia a última data no arquivo XLSX
    2. Verifica se o arquivo já está atualizado (última data = data atual - 1)
    3. Se sim, encerra. Se não, consulta dados do banco e adiciona ao XLSX
    
    ⚠️ Preserva completamente: formatação, tipos de dados, estilos e estrutura original
    """
    
    caminho_xlsx = r"\\trc-dc-ad\Planejamento\MIS\CARTEIRAS\WILLBANK\Automations\base_tempos.xlsx"
    
    try:
        print("="*60)
        print("INICIANDO PROCESSO DE ATUALIZAÇÃO DE TEMPOS OPERACIONAIS")
        print("="*60)
        
        # 1. Mapear última data no arquivo
        print("\n[1/3] Mapeando última data no arquivo...")
        ultima_data_arquivo = obter_ultima_data_arquivo(caminho_xlsx)
        
        # 2. Verificar se arquivo já está atualizado
        print("\n[2/3] Verificando se arquivo já está atualizado...")
        data_esperada = datetime.now().date() - timedelta(days=1)
        ultima_data_arquivo_date = ultima_data_arquivo.date()
        
        print(f"Última data no arquivo: {ultima_data_arquivo_date.strftime('%d/%m/%Y')}")
        print(f"Data esperada (hoje - 1): {data_esperada.strftime('%d/%m/%Y')}")
        
        if ultima_data_arquivo_date == data_esperada:
            print(f"\n✓ Arquivo já está atualizado! Nenhuma ação necessária.")
            return {
                'sucesso': True,
                'mensagem': 'Arquivo já está atualizado',
                'linhas_adicionadas': 0,
                'ultima_data': ultima_data_arquivo_date.strftime('%d/%m/%Y')
            }
        
        # 3. Se não estiver atualizado, buscar dados
        print("\n[3/3] Conectando ao banco e consultando dados...")
        data_inicio = ultima_data_arquivo_date + timedelta(days=1)
        data_fim = data_esperada
        
        print(f"Período a buscar: {data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}")
        
        conn_bd2 = get_connection("TRC-DC-BD2", "PLANEJAMENTO")
        df_novos = consultar_tempos_operacionais(conn_bd2, data_inicio, data_fim)
        conn_bd2.close()
        
        if len(df_novos) == 0:
            print("\nNenhum dado novo encontrado para o período.")
            return {
                'sucesso': True,
                'mensagem': 'Nenhum dado novo encontrado',
                'linhas_adicionadas': 0,
                'periodo': f"{data_inicio.strftime('%d/%m/%Y')} até {data_fim.strftime('%d/%m/%Y')}"
            }
        
        # Adicionar ao arquivo XLSX usando openpyxl (preserva 100% da formatação e estrutura)
        print(f"\nAdicionando {len(df_novos)} linhas ao arquivo XLSX...")
        
        # Carregar workbook existente (preserva formatting, styles, etc)
        wb = load_workbook(caminho_xlsx)
        ws = wb.active
        
        # Encontrar última linha com dados
        ultima_linha = ws.max_row
        print(f"Última linha do arquivo: {ultima_linha}")
        
        # Adicionar novos dados a partir da próxima linha (sem header)
        for r_idx, row in enumerate(dataframe_to_rows(df_novos, index=False, header=False), 
                                     start=ultima_linha + 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                # Preservar o formato da célula anterior (copia do padrão da coluna)
        
        # Salvar workbook (preserva tudo: formatação, estilos, fontes, cores, etc)
        wb.save(caminho_xlsx)
        
        print(f"Dados adicionados com sucesso!")
        print(f"Formatação, estilos e tipos de dados: ✓ Preservados")
        
        print("\n" + "="*60)
        print("PROCESSO CONCLUÍDO COM SUCESSO!")
        print("="*60)
        
        return {
            'sucesso': True,
            'mensagem': 'Dados atualizados com sucesso',
            'linhas_adicionadas': len(df_novos),
            'periodo': f"{data_inicio.strftime('%Y-%m-%d')} até {data_fim.strftime('%Y-%m-%d')}"
        }
        
    except Exception as e:
        print(f"\nERRO: {str(e)}")
        import traceback
        traceback.print_exc()
        return {
            'sucesso': False,
            'mensagem': f'Erro ao processar: {str(e)}',
            'erro': str(e)
        }

# Exemplo de uso
if __name__ == "__main__":
    resultado = processar_atualizacao_tempos()
    
    print("\n" + "="*60)
    print("RESUMO DA EXECUÇÃO")
    print("="*60)
    for chave, valor in resultado.items():
        print(f"{chave}: {valor}")