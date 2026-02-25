import pandas as pd
from openpyxl import load_workbook
import sys
import os
from Getnet.src.consolidacao.data_loader_pipeline import executar_pipeline_completo_com_carregamento
from Ouze.src.config import FILTROS_SQL, DATASETS_TO_LOAD

root_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))

if root_path not in sys.path:
    sys.path.append(root_path)
# ============================================
# NOVO: Função integrada com carregamento
# ============================================
def atualizar_arquivo_funil_completo(caminho_destino=None, datasets_to_load=None):
    """
    ✅ NOVO: Executa o pipeline COMPLETO (carregamento + processamento) 
    e atualiza o arquivo Excel com novos dados.
    
    Esta é a forma RECOMENDADA de atualizar o funil.
    Carrega dados do banco → Processa com arquitetura modular → Salva no Excel.
    
    Args:
        caminho_destino (str, optional): Caminho do arquivo Excel de destino.
            Se None, usa o caminho padrão.
        **filtros_sql: Filtros SQL customizados para cada dataset:
            - where_campanhas
            - where_clientes_mailing
            - where_acionamentos
            - where_tabulacao
            - where_clientes_pagamentos
            - where_clientes_acordos
            - where_massivos
            - where_telefones
    
    Returns:
        dict: Dicionário com informações sobre a execução:
            - 'sucesso': bool indicando se a operação foi bem-sucedida
            - 'mensagem': str com mensagem descritiva
            - 'registros_novos': int com quantidade de registros adicionados
            - 'ultima_data': str com a última data processada
    
    Exemplo de uso:
        # 1. Executa tudo (recomendado)
        resultado = atualizar_arquivo_funil_completo()
        
        # 2. Com filtros específicos
        resultado = atualizar_arquivo_funil_completo(
            where_campanhas="AND A.GrupoPrincipal = 5",
            where_clientes_mailing="COD_CLI = 253"
        )
    """
    
    if datasets_to_load is None:
        datasets_to_load = DATASETS_TO_LOAD

    if caminho_destino is None:
        caminho_destino = r"\\trc-dc-ad\Planejamento\MIS\Pipelines\Data\df_csvBI_padronizado_teste.xlsx"
    
    resultado = {
        'sucesso': False,
        'mensagem': '',
        'registros_novos': 0,
        'ultima_data': None
    }
    
    try:
        # 1. Executa pipeline COMPLETO (carrega dados + processa tudo)
        print("🔄 Executando pipeline completo (carregamento + processamento)...")
        resultados_pipeline = executar_pipeline_completo_com_carregamento(datasets_to_load=datasets_to_load, **FILTROS_SQL)
        
        df_funil_final = resultados_pipeline['consolidado']
        
        # Normalizar nome da coluna de data
        if 'DATA' in df_funil_final.columns:
            df_funil_final.rename(columns={'DATA': 'data'}, inplace=True)
        
        df_funil_final['data'] = pd.to_datetime(df_funil_final['data'], errors='coerce')
        
        try:
            # 2. Lê o arquivo existente
            print("📂 Lendo arquivo existente...")
            df_existente = pd.read_excel(caminho_destino, sheet_name='df_csvBI_padronizado_teste')
            df_existente['data'] = pd.to_datetime(df_existente['data'], errors='coerce')
            
            # 3. Obtém última data
            ultima_data_existente = df_existente['data'].max()
            print(f"📅 Última data no arquivo: {ultima_data_existente.strftime('%Y-%m-%d')}")
            resultado['ultima_data'] = ultima_data_existente.strftime('%Y-%m-%d')
            
            # 4. Filtra apenas novos dados
            df_novos = df_funil_final[df_funil_final['data'] > ultima_data_existente]
            
            if len(df_novos) == 0:
                resultado['sucesso'] = True
                resultado['mensagem'] = "Nenhum dado novo para adicionar."
                resultado['registros_novos'] = 0
                print("⚠️ " + resultado['mensagem'])
                return resultado
            
            print(f"✅ {len(df_novos)} novos registros serão adicionados")
            
            # 5. Carrega workbook e adiciona dados
            wb = load_workbook(caminho_destino)
            ws = wb['df_csvBI_padronizado_teste']
            proxima_linha = ws.max_row + 1
            linha_inicial = proxima_linha
            
            for _, row in df_novos.iterrows():
                for col_idx, valor in enumerate(row, start=1):
                    ws.cell(row=proxima_linha, column=col_idx, value=valor)
                proxima_linha += 1
            
            # 6. Salva arquivo
            wb.save(caminho_destino)
            wb.close()
            
            resultado['sucesso'] = True
            resultado['mensagem'] = f"Arquivo atualizado! {len(df_novos)} registros adicionados."
            resultado['registros_novos'] = len(df_novos)
            
            print(f"💾 {resultado['mensagem']}")
            
        except FileNotFoundError:
            print("⚠️ Arquivo não encontrado. Criando novo...")
            with pd.ExcelWriter(caminho_destino, engine='openpyxl') as writer:
                df_funil_final.to_excel(writer, sheet_name='df_csvBI_padronizado_teste', index=False)
            
            resultado['sucesso'] = True
            resultado['mensagem'] = f"Novo arquivo criado com {len(df_funil_final)} registros."
            resultado['registros_novos'] = len(df_funil_final)
            
            print(f"✅ {resultado['mensagem']}")
        
        return resultado
    
    except Exception as e:
        resultado['sucesso'] = False
        resultado['mensagem'] = f"Erro ao atualizar arquivo: {str(e)}"
        print(f"❌ {resultado['mensagem']}")
        raise


# ============================================
# ORIGINAL: Função legada (preservada)
# ============================================
def atualizar_arquivo_funil(caminho_destino=None):
    """
    Executa o pipeline do funil e atualiza o arquivo Excel com novos dados.
    
    Args:
        caminho_destino (str, optional): Caminho do arquivo Excel de destino.
            Se None, usa o caminho padrão.
    
    Returns:
        dict: Dicionário com informações sobre a execução:
            - 'sucesso': bool indicando se a operação foi bem-sucedida
            - 'mensagem': str com mensagem descritiva
            - 'registros_novos': int com quantidade de registros adicionados
            - 'ultima_data': str com a última data processada
    """
    from pipeline_funil import executar_pipeline_funil
    # Define o caminho padrão se não for fornecido
    if caminho_destino is None:
        caminho_destino = r"\\trc-dc-ad\Planejamento\MIS\Pipelines\df_csvBI_padronizado_teste.xlsx"
    
    resultado = {
        'sucesso': False,
        'mensagem': '',
        'registros_novos': 0,
        'ultima_data': None
    }
    
    try:
        # 1. Executa o pipeline do funil
        print("🔄 Executando pipeline do funil...")
        df_funil_final = executar_pipeline_funil()
        
        # Renomeia 'DATA' para 'data' no df_funil_final para padronizar
        if 'DATA' in df_funil_final.columns:
            df_funil_final.rename(columns={'DATA': 'data'}, inplace=True)
        
        df_funil_final['data'] = pd.to_datetime(df_funil_final['data'], errors='coerce')
        
        try:
            # 2. Lê o arquivo existente da aba 'df_csvBI_padronizado'
            print("📂 Lendo arquivo existente...")
            df_existente = pd.read_excel(caminho_destino, sheet_name='df_csvBI_padronizado_teste')
            
            # 3. Converte a coluna 'data' para datetime
            df_existente['data'] = pd.to_datetime(df_existente['data'], errors='coerce')
            
            # 4. Obtém a última data do arquivo existente
            ultima_data_existente = df_existente['data'].max()
            
            print(f"📅 Última data no arquivo: {ultima_data_existente.strftime('%Y-%m-%d')}")
            resultado['ultima_data'] = ultima_data_existente.strftime('%Y-%m-%d')
            
            # 5. Filtra apenas dados NOVOS (posteriores à última data existente)
            df_novos = df_funil_final[df_funil_final['data'] > ultima_data_existente]
            
            if len(df_novos) == 0:
                resultado['sucesso'] = True
                resultado['mensagem'] = "Nenhum dado novo para adicionar. Arquivo não foi atualizado."
                resultado['registros_novos'] = 0
                print("⚠️ " + resultado['mensagem'])
                return resultado
            
            print(f"✅ {len(df_novos)} novos registros serão adicionados")
            
            # 6. Carrega o workbook existente SEM destruir formatação
            wb = load_workbook(caminho_destino)
            ws = wb['df_csvBI_padronizado_teste']
            
            # 7. Encontra a próxima linha vazia (após a última linha com dados)
            proxima_linha = ws.max_row + 1
            linha_inicial = proxima_linha
            
            # 8. Adiciona apenas os DADOS NOVOS (sem cabeçalho)
            for _, row in df_novos.iterrows():
                for col_idx, valor in enumerate(row, start=1):
                    ws.cell(row=proxima_linha, column=col_idx, value=valor)
                proxima_linha += 1
            
            # 9. Salva o arquivo preservando TUDO (formatação, abas, tabelas dinâmicas)
            wb.save(caminho_destino)
            wb.close()
            
            resultado['sucesso'] = True
            resultado['mensagem'] = f"Arquivo atualizado com sucesso! {len(df_novos)} registros adicionados."
            resultado['registros_novos'] = len(df_novos)
            
            print(f"💾 {resultado['mensagem']}")
            print(f"📊 Registros adicionados a partir da linha {linha_inicial}")
            
        except FileNotFoundError:
            # Arquivo não existe, cria um novo
            print("⚠️ Arquivo não encontrado. Criando novo arquivo com os dados gerados.")
            with pd.ExcelWriter(caminho_destino, engine='openpyxl') as writer:
                df_funil_final.to_excel(writer, sheet_name='df_csvBI_padronizado_teste', index=False)
            
            resultado['sucesso'] = True
            resultado['mensagem'] = "Arquivo criado com sucesso!"
            resultado['registros_novos'] = len(df_funil_final)
            resultado['ultima_data'] = df_funil_final['data'].max().strftime('%Y-%m-%d')
            
            print(f"💾 {resultado['mensagem']}")
        
        return resultado
        
    except Exception as e:
        resultado['sucesso'] = False
        resultado['mensagem'] = f"Erro ao processar arquivo: {str(e)}"
        print(f"❌ {resultado['mensagem']}")
        raise

# Exemplo de uso:
if __name__ == "__main__":
    print("\n" + "="*80)
    print("PIPELINE DE FUNIL GETNET")
    print("="*80 + "\n")
    
    # ============================================
    # OPÇÃO 1: NOVO - Pipeline completo (RECOMENDADO)
    # ============================================
    print("Executando: atualizar_arquivo_funil_completo() [RECOMENDADO]")
    print("(Carregamento + Processamento Modular + Consolidação)\n")
    
    resultado = atualizar_arquivo_funil_completo(datasets_to_load=DATASETS_TO_LOAD)
    
    print("\n📋 Resultado da execução:")
    print(f"   ✅ Sucesso: {resultado['sucesso']}")
    print(f"   📝 Mensagem: {resultado['mensagem']}")
    print(f"   📊 Novos registros: {resultado['registros_novos']}")
    print(f"   📅 Última data: {resultado['ultima_data']}")
    
    # ============================================
    # OPÇÃO 2: ORIGINAL - Pipeline legado (preservado)
    # ============================================
    # Descomente para usar a função antiga:
    # resultado = atualizar_arquivo_funil()
    
    print("\n" + "="*80)
    if resultado['sucesso']:
        print("✅ EXECUÇÃO CONCLUÍDA COM SUCESSO")
    else:
        print("❌ EXECUÇÃO FALHOU")
    print("="*80 + "\n")