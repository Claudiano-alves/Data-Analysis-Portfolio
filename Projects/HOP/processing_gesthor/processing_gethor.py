import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIGURAÇÕES ─────────────────────────────────────────────────────────────
PASTA_ORIGEM = r"\\trc-dc-ad\Grades\HOP\Relatorios\Planilhão - relatorios de agendamentos\Gesthor diario\2026"
ARQUIVO_SAIDA = Path(__file__).parent / "CONSOLIDADO_2026.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

COLUNAS_ORIGEM = list("ABCDEFGHIJKL")  # A até L (12 colunas)

ROTULOS_DESTINO = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "Agendamento"
]

def ler_arquivo(caminho: Path) -> pd.DataFrame | None:
    """Lê colunas A:L de um arquivo xlsx, ignorando cabeçalho."""
    try:
        df = pd.read_excel(
            caminho,
            usecols="A:L",
            header=0,        # primeira linha como cabeçalho (será descartada no rename)
            dtype=str,
        )
        # Garante exatamente 12 colunas
        if df.shape[1] < 12:
            print(f"  [AVISO] {caminho.name} tem menos de 12 colunas — ignorado.")
            return None

        df = df.iloc[:, :12]  # pega só as 12 primeiras caso haja mais
        df.columns = ROTULOS_DESTINO

        # Remove linhas completamente vazias
        df.dropna(how="all", inplace=True)

        # Adiciona coluna de origem para rastreabilidade
        df.insert(0, "Arquivo_Origem", caminho.name)

        return df

    except Exception as e:
        print(f"  [ERRO] Não foi possível ler {caminho.name}: {e}")
        return None


def consolidar():
    pasta = Path(PASTA_ORIGEM)

    if not pasta.exists():
        print(f"[ERRO] Pasta não encontrada: {PASTA_ORIGEM}")
        return

    arquivos = sorted([
        f for f in pasta.glob("*.xlsx")
        if f.name != ARQUIVO_SAIDA.name  # evita ler o próprio consolidado
    ])

    if not arquivos:
        print("[AVISO] Nenhum arquivo .xlsx encontrado na pasta.")
        return

    print(f"Encontrados {len(arquivos)} arquivo(s). Consolidando...\n")

    frames = []
    for arq in arquivos:
        print(f"  Lendo: {arq.name}")
        df = ler_arquivo(arq)
        if df is not None:
            frames.append(df)

    if not frames:
        print("\n[ERRO] Nenhum dado válido encontrado.")
        return

    consolidado = pd.concat(frames, ignore_index=True)

    # ── Salva com openpyxl para aplicar formatação ────────────────────────────
    consolidado.to_excel(ARQUIVO_SAIDA, index=False, engine="openpyxl")

    # ── Formatação profissional ───────────────────────────────────────────────
    wb = load_workbook(ARQUIVO_SAIDA)
    ws = wb.active
    ws.title = "Consolidado"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border_side = Side(style="thin", color="CCCCCC")
    cell_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )

    # Cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = cell_border

    ws.row_dimensions[1].height = 30

    # Dados
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_align
            cell.border = cell_border

    # Larguras de coluna automáticas
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Filtros automáticos
    ws.auto_filter.ref = ws.dimensions

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    wb.save(ARQUIVO_SAIDA)

    total_linhas = len(consolidado)
    print(f"\n✅ Consolidação concluída!")
    print(f"   Arquivos processados : {len(frames)}")
    print(f"   Total de linhas      : {total_linhas:,}")
    print(f"   Arquivo salvo em     : {ARQUIVO_SAIDA}")


if __name__ == "__main__":
    consolidar()