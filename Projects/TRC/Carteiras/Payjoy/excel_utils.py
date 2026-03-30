import pandas as pd
from dotenv import load_dotenv
from db_connection import get_connection
from datetime import datetime, timedelta
from glob import glob
import os
import sys
import re
import logging
from openpyxl import load_workbook

xlsx_file = r"\\trc-dc-ad\Planejamento\MIS\CARTEIRAS\PayJoy\base_PAYJOY.xlsx"
# Logging configuration: use INFO by default; set PAYJOY_DEBUG=1 to see debug messages
log_level = logging.DEBUG if os.getenv('PAYJOY_DEBUG', '0') == '1' else logging.INFO
logging.basicConfig(level=log_level, format='[%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)


def atualizar_arquivo_excel_por_df(xlsx_file, df_payjoy):
    """
    Atualiza o arquivo Excel (`xlsx_file`) com as colunas de data presentes em
    `df_payjoy` (formato wide: FAIXA | Indicador | 16/12 | 17/12 | ...).

    Regras:
      - Só adiciona datas **após** a última data já presente no arquivo.
      - Usa o ano da última data existente no arquivo para interpretar colunas do
        df_payjoy no formato 'DD/MM'.

    Retorna um dicionário resumo: {'updated': bool, 'colunas_adicionadas': [...], 'dados_inseridos': int}
    """
    if df_payjoy is None or df_payjoy.empty:
        logger.info("Nenhum dado em df_payjoy; nada a atualizar")
        return {'updated': False, 'reason': 'df_empty'}

    # Carrega arquivo e mapeia datas existentes
    try:
        wb = load_workbook(xlsx_file)
        ws = wb.active
        df_existente = pd.read_excel(xlsx_file)
    except Exception as e:
        logger.error(f"Erro ao abrir arquivo Excel: {e}")
        return {'updated': False, 'error': str(e)}

    colunas_fixas = ['FAIXA', 'Indicador']

    def _col_to_date_obj(col, default_year):
        if isinstance(col, (pd.Timestamp, datetime)):
            d = col.to_pydatetime() if isinstance(col, pd.Timestamp) else col
            return d.date()
        s = str(col).strip()
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y']:
            try:
                return datetime.strptime(s, fmt).date()
            except:
                continue
        # suporte para 'DD/MM' sem ano -> usa default_year
        if '/' in s:
            try:
                dia, mes = s.split('/')[:2]
                return datetime(default_year, int(mes), int(dia)).date()
            except:
                pass
        return None

    # Existing dates
    existing_date_positions = {}
    colunas_existentes = df_existente.columns.tolist()
    for idx, col in enumerate(colunas_existentes, start=1):
        if col in colunas_fixas:
            continue
        date_obj = _col_to_date_obj(col, datetime.now().year)
        if date_obj:
            existing_date_positions[date_obj] = idx

    if existing_date_positions:
        last_existing_date = max(existing_date_positions.keys())
        logger.info(f"Última data existente no Excel: {last_existing_date.strftime('%Y-%m-%d')}")
    else:
        # Se não houver datas, usamos o ano atual como referência
        last_existing_date = datetime.now().date() - timedelta(days=1)
        logger.info("Nenhuma data existente encontrada no Excel; usando ano atual como referência")

    # Extrai datas do df_payjoy
    df_pay_dates = {}
    for col in df_payjoy.columns:
        if col in colunas_fixas:
            continue
        date_obj = _col_to_date_obj(col, last_existing_date.year)
        if date_obj:
            df_pay_dates[col] = date_obj
        else:
            logger.debug(f"Coluna do df_payjoy não reconhecida como data: {col}")

    # Determina quais datas são posteriores à última data do arquivo
    to_add = sorted([d for d in set(df_pay_dates.values()) if d > last_existing_date])

    if not to_add:
        logger.info("Nenhuma data nova após a última data do arquivo; nada a adicionar")
        return {'updated': False, 'colunas_adicionadas': [], 'dados_inseridos': 0}

    # Inserir colunas em posições corretas (ordenadas por data)
    logger.info(f"Adicionando {len(to_add)} novas datas ao arquivo: {[d.strftime('%Y-%m-%d') for d in to_add]}")
    # Reconstroi mapping para posições dinâmicas
    existing_positions = dict(existing_date_positions)
    colunas_existentes = df_existente.columns.tolist()

    for d in to_add:
        # encontra primeira existing date maior que d
        insert_before_idx = None
        for exist_date, pos in sorted(existing_positions.items(), key=lambda x: x[1]):
            if exist_date > d:
                insert_before_idx = pos
                break
        insert_idx = insert_before_idx if insert_before_idx is not None else len(colunas_existentes) + 1

        ws.insert_cols(insert_idx)
        ws.cell(row=1, column=insert_idx, value=datetime.combine(d, datetime.min.time()))
        logger.debug(f"Inserida coluna para {d.strftime('%Y-%m-%d')} na posição {insert_idx}")

        # atualizar positions
        updated = {}
        for exist_date, pos in existing_positions.items():
            updated[exist_date] = pos + 1 if pos >= insert_idx else pos
        existing_positions = updated
        existing_positions[d] = insert_idx
        colunas_existentes.insert(insert_idx - 1, d)

    # Reconstrói map date->colidx
    date_to_colidx = {date: idx for date, idx in existing_positions.items()}

    # Mapeamento FAIXA+Indicador -> linha
    faixa_indicador_para_linha = {}
    for idx in range(2, ws.max_row + 1):
        faixa = ws.cell(row=idx, column=1).value
        indicador = ws.cell(row=idx, column=2).value
        if faixa is not None and indicador is not None:
            chave = (str(faixa).strip(), str(indicador).strip())
            faixa_indicador_para_linha[chave] = idx

    # Insere valores
    dados_inseridos = 0
    for _, row in df_payjoy.iterrows():
        faixa = str(row['FAIXA']).strip()
        indicador = str(row['Indicador']).strip()
        chave = (faixa, indicador)
        if chave not in faixa_indicador_para_linha:
            continue
        linha_excel = faixa_indicador_para_linha[chave]
        for col_name, date_obj in df_pay_dates.items():
            if date_obj in date_to_colidx:
                col_idx = date_to_colidx[date_obj]
                valor = row[col_name]
                if pd.notna(valor):
                    ws.cell(row=linha_excel, column=col_idx, value=valor)
                    dados_inseridos += 1

    wb.save(xlsx_file)
    logger.info(f"Arquivo salvo com {len(to_add)} colunas adicionadas e {dados_inseridos} valores inseridos")

    return {'updated': True, 'colunas_adicionadas': [d.strftime('%Y-%m-%d') for d in to_add], 'dados_inseridos': dados_inseridos}

def identificar_maior_data_xlsx(xlsx_path):
    """
    Identifica a maior data presente nas colunas do arquivo Excel.
    Encerra a execução se o arquivo não existir.
    """
    if not os.path.exists(xlsx_path):
        print(f"\n{'='*80}")
        print(f"ERRO CRÍTICO: Arquivo não encontrado!")
        print(f"Caminho: {xlsx_path}")
        print(f"{'='*80}")
        print("\nEncerrando execução...")
        sys.exit(1)
    
    try:
        df_temp = pd.read_excel(xlsx_path, nrows=5)
        logger.info("Arquivo Excel lido com sucesso")
        
        colunas = df_temp.columns.tolist()
        logger.debug(f"Colunas encontradas no arquivo: {colunas}")
        
        ultima_data = None
        formatos = ['%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d-%m-%Y']
        
        for col in reversed(colunas):
            if col in ['FAIXA', 'Indicador']:
                continue
            
            try:
                # Se já é datetime/Timestamp, converter e usar
                if isinstance(col, (pd.Timestamp, datetime)):
                    if isinstance(col, pd.Timestamp):
                        ultima_data = col.to_pydatetime().date()
                    else:
                        ultima_data = col.date() if hasattr(col, 'date') else col
                    logger.info(f"Maior data encontrada: {ultima_data.strftime('%Y-%m-%d')} (coluna já era datetime)")
                    print(f"\nMaior data encontrada: {ultima_data.strftime('%Y-%m-%d')} (coluna: '{col}')")
                    return ultima_data
                
                # Se é string, tentar múltiplos formatos
                col_str = str(col).strip()
                for formato in formatos:
                    try:
                        data = datetime.strptime(col_str, formato).date()
                        ultima_data = data
                        logger.info(f"Maior data encontrada: {ultima_data.strftime('%Y-%m-%d')} (formato: {formato})")
                        print(f"\nMaior data encontrada: {ultima_data.strftime('%Y-%m-%d')} (coluna: '{col}')")
                        return ultima_data
                    except ValueError:
                        continue
                        
            except Exception as e:
                logger.debug(f"Coluna '{col}' não é data válida: {e}")
                continue
        
        if ultima_data is None:
            print("\n" + "="*80)
            print("ERRO: Nenhuma data válida encontrada nas colunas do arquivo!")
            print("Colunas analisadas:", [col for col in colunas if col not in ['FAIXA', 'Indicador']])
            print("="*80)
            print("\nEncerrando execução...")
            sys.exit(1)
        
        return ultima_data
            
    except Exception as e:
        print(f"\n{'='*80}")
        print(f"ERRO ao ler arquivo Excel: {e}")
        print(f"{'='*80}")
        import traceback
        traceback.print_exc()
        print("\nEncerrando execução...")
        sys.exit(1)